"""Generate Data/20_scenarios.xlsx -- the PROJECT-WIDE scenario selection.

One file, read by every model, so that choosing a scenario once applies it
everywhere. Previously this lived in 19_ sheet Scenarios, which is
ADAS-specific and the wrong home for a project-wide switch.

    python3 tools/make_20_scenarios.py

Reasoning: docs/04_SENSOR_MODEL_DESIGN.md section 7, and
docs/05_ADAS_SENSOR_ADOPTION_REPORT.md sections 7.2 and 7.3.

Colour convention, same as 17_ / 18_ / 19_:
    YELLOW = editable input
    ORANGE = assumption with no source behind it
    GREEN  = fact, sourced and dated
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "Data" / "20_scenarios.xlsx"

YELLOW = PatternFill("solid", fgColor="FFF2CC")
ORANGE = PatternFill("solid", fgColor="FCE4D6")
GREY = PatternFill("solid", fgColor="EDEDED")
HDR = PatternFill("solid", fgColor="D9D9D9")

BOLD = Font(bold=True)
TITLE = Font(bold=True, size=13)
BIG = Font(bold=True, size=12)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

NOTE_COL = 12          # free-text column, clear of every data column
SCENARIO_YEARS = [2040, 2050, 2060, 2070]

# Driver C -- multiplier on SENSOR COUNTS after 2040. All scenarios are exactly
# 1.00 at 2040 by construction, so this cannot perturb any year for which
# sourced data exists (validation V9).
SCENARIOS = [
    ("S1", "Cost-driven convergence", 1.00, 1.00, 1.00, 1.00, 0.25,
     "Sensors get cheap, but sensing converges on a minimal efficient set "
     "(camera + radar). Redundancy achieved in COMPUTE, not hardware. The "
     "vision-only thesis."),
    ("S2", "Regulated redundancy", 1.00, 1.15, 1.28, 1.40, 0.50,
     "Cost decline makes redundancy affordable; L3/L4 type approval requires "
     "independent modality redundancy for fail-operational behaviour. MODE."),
    ("S3", "Full fail-operational", 1.00, 1.35, 1.70, 2.00, 0.25,
     "Private L4 requires dual-redundant sensing across every modality plus "
     "degraded-mode operation. Each sensing function duplicated."),
]

# How each model is expected to read a scenario. NOT low/medium/high --
# see the Notes sheet.
DIRECTION = [
    ("Sensor count", "lowest", "mid", "highest",
     "S1 converges on a minimal sensor set."),
    ("Wiring length", "lowest", "mid", "highest",
     "Follows sensor count, via metres-per-sensor."),
    ("PCB / compute area", "HIGHEST", "mid", "highest",
     "S1 moves redundancy INTO silicon -- fewer sensors, more compute. "
     "ANTI-CORRELATED with the two rows above."),
]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = HDR
        cell.border = THIN
        cell.alignment = WRAP


def widths(ws, *pairs):
    for col, w in pairs:
        ws.column_dimensions[col].width = w


def build():
    wb = Workbook()

    # ---------------- Control ----------------
    ws = wb.active
    ws.title = "Control"
    ws["A1"] = "PROJECT-WIDE SCENARIO SELECTION"
    ws["A1"].font = TITLE
    ws["A2"] = ("This one cell drives EVERY model -- wiring, sensors, PCB. "
                "Change it here and nowhere else.")
    ws["A2"].alignment = WRAP

    ws.cell(row=4, column=1, value="Active_Scenario").font = BIG
    c = ws.cell(row=4, column=2, value="SAMPLE")
    c.fill = YELLOW
    c.border = THIN
    c.font = BIG

    ws.cell(row=6, column=1, value="Allowed values").font = BOLD
    rows = [
        ("SAMPLE", "DEFAULT. Each Monte Carlo iteration draws a scenario by "
                   "weight, so ONE run's band spans all three. This is what "
                   "should be quoted -- pinning discards the scenario spread "
                   "and reports a narrower band than the evidence supports."),
        ("S1", "Pin every iteration to S1. Outputs take a _S1 suffix."),
        ("S2", "Pin every iteration to S2. Outputs take a _S2 suffix."),
        ("S3", "Pin every iteration to S3. Outputs take a _S3 suffix."),
    ]
    for i, (k, v) in enumerate(rows, start=7):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v).alignment = WRAP

    ws.cell(row=12, column=1, value="Code override").font = BOLD
    ws.cell(row=12, column=2,
            value="SCENARIO_OVERRIDE in BevWiring.py section 0 beats this cell, "
                  "for scripted sweeps across all three without editing Excel. "
                  "Set it to None to obey this file.").alignment = WRAP
    ws.cell(row=14, column=1, value="Read by").font = BOLD
    ws.cell(row=14, column=2,
            value="Wiring/BevWiring.py. SensorNumbersMC.py and the PCB models "
                  "once they are rewired -- see docs/04_SENSOR_MODEL_DESIGN.md "
                  "section 9.").alignment = WRAP
    widths(ws, ("A", 20), ("B", 105))

    # ---------------- Scenarios ----------------
    ws = wb.create_sheet("Scenarios")
    ws["A1"] = "Driver C -- post-2040 sensor-count scenarios"
    ws["A1"].font = TITLE
    ws["A2"] = (
        "A model needs to know HOW MANY sensors, not WHICH. After 2040 the "
        "count is set by two observable mechanisms: (1) cost decline with "
        "volume -- by ~2040-45 sensor cost is negligible against vehicle "
        "price, so cost stops constraining count; (2) safety redundancy -- "
        "once cost is not binding, required redundancy sets the count. A new "
        "sensor modality, if one arrives, simply appears as 'more sensors'. "
        "This replaces the withdrawn rule E4 ('no new modality before 2070'), "
        "which assumed zero over 45 years against a historical rate of roughly "
        "one new modality every 10-15 years.")
    ws["A2"].alignment = WRAP

    hdr = ["Scenario", "Name"] + [str(y) for y in SCENARIO_YEARS] + \
          ["Weight", "Mechanism"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, len(hdr))
    for i, row in enumerate(SCENARIOS, start=5):
        ws.cell(row=i, column=1, value=row[0]).font = BOLD
        ws.cell(row=i, column=2, value=row[1])
        for j, v in enumerate(row[2:6], start=3):
            c = ws.cell(row=i, column=j, value=v)
            c.fill = YELLOW
            c.border = THIN
            c.number_format = "0.00"
        c = ws.cell(row=i, column=7, value=row[6])
        c.fill = YELLOW
        c.border = THIN
        c.number_format = "0.00"
        ws.cell(row=i, column=8, value=row[7]).alignment = WRAP

    r = 5 + len(SCENARIOS)
    ws.cell(row=r, column=2, value="Weights sum (V10, must be 1.000)").font = BOLD
    c = ws.cell(row=r, column=7, value=f"=SUM(G5:G{r - 1})")
    c.number_format = "0.000"
    c.fill = GREY

    ws.cell(row=r + 2, column=NOTE_COL,
            value="All scenarios are 1.00 at 2040 BY CONSTRUCTION, so Driver C "
                  "cannot perturb any year for which sourced data exists "
                  "(validation V9). Anchors are read through PCHIP -- add or "
                  "delete year columns freely.").alignment = WRAP
    widths(ws, ("A", 11), ("B", 26), ("C", 8), ("D", 8), ("E", 8), ("F", 8),
           ("G", 9), ("H", 60), (get_column_letter(NOTE_COL), 70))

    # ---------------- Notes ----------------
    ws = wb.create_sheet("Notes")
    ws["A1"] = "READ BEFORE WIRING A NEW MODEL TO THIS FILE"
    ws["A1"].font = TITLE

    ws.cell(row=3, column=1,
            value="SCENARIOS ARE NOT LOW / MEDIUM / HIGH").font = BIG
    ws.cell(row=4, column=1,
            value="S1 says: sensors converge on a minimal set, and redundancy "
                  "is achieved in COMPUTE rather than in hardware. So S1 is the "
                  "LOW case for sensors and wiring but the HIGH case for PCB "
                  "and compute area. Wiring S1 to 'low' everywhere would delete "
                  "that substitution and understate PCB uncertainty.").alignment = WRAP

    hdr = ["Quantity", "S1", "S2", "S3", "Why"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=6, column=j, value=h)
    style_header(ws, 6, len(hdr))
    for i, row in enumerate(DIRECTION, start=7):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN
            if j == 5:
                c.alignment = WRAP
        if row[0].startswith("PCB"):
            ws.cell(row=i, column=2).fill = ORANGE
            ws.cell(row=i, column=2).font = BOLD

    notes = [
        ("Same trap, three times",
         "This is the same structural problem as CAN-vs-Ethernet in "
         "BevWiring_STATUS.md section 10, and as Presence_per_Tier row 9 in "
         "19_, where 'ADAS camera ECU (basic)' falls 1.00 -> 0.10 as the "
         "domain controller and central computer rise. Components that "
         "SUBSTITUTE for each other must be drawn jointly, never "
         "independently."),
        ("Why SAMPLE is the default",
         "Pinning to one scenario produces a narrower band than the evidence "
         "supports, because it silently discards the S1/S3 spread. Pinned runs "
         "are for UNDERSTANDING the scenarios; SAMPLE is what to quote."),
        ("Weakest input in the chain",
         "The multipliers 1.0 / 1.4 / 2.0 are the largest single lever on the "
         "2070 answer and are NOT sourced. The two mechanisms behind them are "
         "observable; the specific values are judgment. See "
         "docs/05_ADAS_SENSOR_ADOPTION_REPORT.md section 9.2."),
        ("History",
         "Created 2026-08-05, moved out of 19_ sheet Scenarios so that one "
         "selection drives every model. 19_ keeps tiers, lidar and presence."),
    ]
    r = 7 + len(DIRECTION) + 2
    for k, v in notes:
        ws.cell(row=r, column=1, value=k).font = BOLD
        ws.cell(row=r, column=2, value=v).alignment = WRAP
        r += 2
    widths(ws, ("A", 26), ("B", 100), ("C", 10), ("D", 10), ("E", 46))

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print("written:", build())
