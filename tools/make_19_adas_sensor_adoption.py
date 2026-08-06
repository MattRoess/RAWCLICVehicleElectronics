"""Generate Data/19_ADAS_sensor_adoption.xlsx from the report.

Source of every number: docs/ADAS_Sensor_Adoption_Report_2025_2070.md
Section references in the sheets point back to it. Regenerate with:

    python3 tools/make_19_adas_sensor_adoption.py

Colour convention, same as 17_ and 18_:
    YELLOW = editable input
    ORANGE = assumption with no source behind it
    GREEN  = fact, sourced and dated
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# This script lives in tools/; its output belongs in Data/ with the other model
# inputs, and its source document in docs/. Paths are anchored on the repo root
# so the script runs from anywhere.
REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "Data" / "19_ADAS_sensor_adoption.xlsx"
REPORT = "docs/ADAS_Sensor_Adoption_Report_2025_2070.md"

YELLOW = PatternFill("solid", fgColor="FFF2CC")
ORANGE = PatternFill("solid", fgColor="FCE4D6")
GREEN = PatternFill("solid", fgColor="E2EFDA")
GREY = PatternFill("solid", fgColor="EDEDED")
HDR = PatternFill("solid", fgColor="D9D9D9")

BOLD = Font(bold=True)
TITLE = Font(bold=True, size=13)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# Column used for free-text footnotes on data sheets. Kept clear of every data
# column so a "column is not null" filter cannot pick notes up as records.
NOTE_COL = 15   # column O

TIERS = ["H0", "H1", "H2", "H3", "H4"]
SEGMENTS = ["AB", "CD", "EF"]
YEARS = [2025, 2030, 2035, 2040, 2050, 2060, 2070]

# --- report section 3 -------------------------------------------------------
TIER_DEF = [
    # tier, name, cam_min, cam_max, rad_min, rad_max, us_min, us_max,
    # lid_min, lid_max, ranging, sae, tag
    ("H0", "GSR-2 mandated floor", 1, 1, 1, 1, 0, 4, 0, 0, 1.0, "L0 / L1", "DERIVED"),
    ("H1", "ACC + lane keeping", 1, 2, 1, 3, 8, 12, 0, 0, 2.0, "L2", "ASSUMPTION"),
    ("H2", "hands-off highway", 5, 8, 3, 5, 12, 12, 0, 0, 4.0, "L2+", "ASSUMPTION"),
    ("H3", "urban + highway 'L2++'", 8, 12, 5, 5, 12, 16, 0, 1, 5.3, "L2++ or L3", "FACT-anchored"),
    ("H4", "redundant, liability transfer", 6, 12, 5, 6, 12, 13, 1, 3, 7.0, "L3", "FACT-anchored"),
]

# --- report section 4.1 -----------------------------------------------------
TIER_SHARES = {
    "EF": {2025: (0.00, 0.10, 0.25, 0.50, 0.15),
           2030: (0.00, 0.05, 0.15, 0.50, 0.30),
           2035: (0.00, 0.00, 0.10, 0.45, 0.45),
           2040: (0.00, 0.00, 0.05, 0.40, 0.55),
           2050: (0.00, 0.00, 0.00, 0.35, 0.65),
           2060: (0.00, 0.00, 0.00, 0.35, 0.65),
           2070: (0.00, 0.00, 0.00, 0.35, 0.65)},
    "CD": {2025: (0.10, 0.45, 0.35, 0.10, 0.00),
           2030: (0.05, 0.25, 0.40, 0.28, 0.02),
           2035: (0.00, 0.10, 0.30, 0.45, 0.15),
           2040: (0.00, 0.05, 0.20, 0.50, 0.25),
           2050: (0.00, 0.00, 0.10, 0.50, 0.40),
           2060: (0.00, 0.00, 0.10, 0.50, 0.40),
           2070: (0.00, 0.00, 0.10, 0.50, 0.40)},
    "AB": {2025: (0.45, 0.45, 0.10, 0.00, 0.00),
           2030: (0.25, 0.45, 0.25, 0.05, 0.00),
           2035: (0.10, 0.35, 0.40, 0.15, 0.00),
           2040: (0.05, 0.25, 0.40, 0.28, 0.02),
           2050: (0.00, 0.10, 0.30, 0.45, 0.15),
           2060: (0.00, 0.10, 0.30, 0.45, 0.15),
           2070: (0.00, 0.10, 0.30, 0.45, 0.15)},
}

# --- report section 5.3 -----------------------------------------------------
LIDAR = [
    (2025, 0.00, 0.01, 0.02, "FACT", "EQS and i7 only; both now withdrawn"),
    (2030, 0.03, 0.12, 0.30, "DERIVED", "Goldman: 3 m overseas ADAS lidar units by 2030"),
    (2035, 0.08, 0.28, 0.55, "ASSUMPTION", "China 21% (2025) + sampled lag"),
    (2040, 0.12, 0.45, 0.70, "ASSUMPTION", ""),
    (2050, 0.18, 0.60, 0.85, "ASSUMPTION", ""),
    (2060, 0.22, 0.66, 0.88, "ASSUMPTION", "E3: rise decelerates"),
    (2070, 0.25, 0.70, 0.90, "ASSUMPTION", "E3: cost curve near materials floor"),
]

# --- report section 6.2 -----------------------------------------------------
# Marks a cell governed by Driver B (the Lidar sheet), not by the tier mix.
# NO LEADING "=" -- Excel reads that as a formula, openpyxl then writes a
# formula cell with no cached value, and every reader gets NaN instead of the
# marker. That silently produced NaN sensor counts once already.
LD = "Driver B"
PRESENCE = [
    ("Front ADAS camera", 1.00, 1.00, 1.00, 1.00, 1.00, "FACT", "GSR-2 mandated since Jul 2024"),
    ("Rear view camera", 0.50, 0.90, 1.00, 1.00, 1.00, "ASSUMPTION", ""),
    ("Side / mirror cameras", 0.00, 0.25, 0.80, 1.00, 1.00, "ASSUMPTION", "360 deg needs >=4 cameras"),
    ("Front long-range radar", 1.00, 1.00, 1.00, 1.00, 1.00, "FACT", "GSR-2 AEB"),
    ("Corner short/mid-range radars", 0.00, 0.30, 0.85, 1.00, 1.00, "FACT-anchored", "5 radar at H3: EQS, EX90"),
    ("LiDAR sensor", 0.00, 0.00, 0.00, LD, LD, "DRIVER B", "NOT tier-governed - see Lidar sheet"),
    ("Ultrasonic sensors", 0.30, 0.90, 1.00, 1.00, 1.00, "ASSUMPTION", ""),
    ("Driver monitoring camera", 0.20, 0.50, 0.90, 1.00, 1.00, "DERIVED", "GSR-2 attention warning"),
    ("ADAS camera ECU (basic)", 1.00, 1.00, 0.50, 0.20, 0.10, "ASSUMPTION", "DECLINES - absorbed by rows 10/11"),
    ("ADAS domain controller / fusion ECU", 0.00, 0.10, 0.70, 1.00, 1.00, "ASSUMPTION", ""),
    ("Automated driving central computer", 0.00, 0.00, 0.10, 0.50, 1.00, "ASSUMPTION", ""),
    ("Parking assist ECU", 0.20, 0.80, 1.00, 1.00, 1.00, "ASSUMPTION", ""),
]

# --- report section 4.2, 5.3 ------------------------------------------------
PARAMS = [
    ("Tier_Band_Shift_Y", 5.0,
     "Min/Max band for Driver A, as a TIME SHIFT on the whole tier curve. "
     "Shifting preserves sum-to-1 by construction; independent Min/Mode/Max "
     "columns per tier would not.", "ASSUMPTION"),
    ("Lidar_Europe_Lag_Y_Mean", 7.0,
     "China reached 21% lidar penetration in NEVs in 2025. Europe follows "
     "after this lag. SAMPLED, not fixed, so the hypothesis is testable.", "ASSUMPTION"),
    ("Lidar_Europe_Lag_Y_SD", 3.0,
     "Spread on the lag above.", "ASSUMPTION"),
    ("Lidar_Segment_Offset_Factor", 0.5,
     "Lidar's segment offset as a fraction of Driver A's. Half, because the "
     "BYD Seagull (A-segment, ~$10,300) shows lidar entering from below.", "ASSUMPTION"),
    ("Lidar_Units_Per_Equipped_H3", 1.0,
     "Measured: EQS 1, i7 1. No European car has shipped with more.", "FACT"),
    ("Lidar_Units_Per_Equipped_H4", 1.0,
     "Mode. Range 1-3; multi-lidar exists only in Chinese flagships.", "ASSUMPTION"),
]

# --- NEW: uncertainty calibrated from observed source disagreement ----------
UNCERTAINTY = [
    ("L2+/L3 share of global new sales, 2035", "S&P Global Mobility", ">=31%",
     "IDTechEx", ">50%", "1.6x",
     "Two top-tier houses, same quantity, same year. Sets the floor for any "
     "10-year-out forecast band."),
    ("Ranging sensors per vehicle, 2035", "Yole", "5.5",
     "this report S4.1", "4.6", "1.20x",
     "Our tier table is deliberately conservative on the late ramp."),
    ("AB->EF adoption lag", "18_ offsets", "10 y",
     "this report S4.1", "15 y", "5 y",
     "Confirms +/-5 y is the right order for the Driver A band."),
    ("CD copper per vehicle, 2025", "model, length-first", "56.4 kg",
     "source report, stated", "60.3 kg", "6.9%",
     "Two readings of the SAME source."),
    ("EF wiring length, 2025", "report row sum", "3646 m",
     "report stated total", "3546 m", "2.8%",
     "A single source disagreeing WITH ITSELF. Sets the noise floor: nothing "
     "here can claim better than ~3%."),
    ("SDV total length reduction", "17_ per-category mechanisms", "-12/-11/-8%",
     "17_ stated segment totals", "-44/-40/-23%", "3-4x",
     "Largest spread in the whole chain. Structural, not statistical."),
    ("EF cameras, 2025", "06_ max", "5",
     "measured EQS / EX90", "6 / 10", "2x",
     "Within normal disagreement - NOT necessarily an error."),
    ("EF lidar presence, 2025", "01_ label 'Opt'", "0.50",
     "measured + Driver B", "0.01", "50x",
     "FAR outside every band above. This one IS an error."),
    ("Europe lidar share, 2070", "this report Min", "0.25",
     "this report Max", "0.90", "3.6x",
     "Genuinely unknown. Width is honest, not lazy."),
]

UNCERTAINTY_LESSONS = [
    ("Self-consistency floor", "~3%",
     "A single source disagrees with itself by 2.8-6.9%. No result from this "
     "chain can claim better precision than that, however many iterations are run."),
    ("Cross-house forecast spread, 10 y out", "~1.6x",
     "S&P vs IDTechEx on the same 2035 quantity. Any single-source forecast "
     "band narrower than this is over-confident."),
    ("Structural / mechanism uncertainty", "3-4x",
     "SDV depth, lidar 2070. Where the MECHANISM is contested rather than the "
     "value, spreads are multiples, not percentages."),
    ("Threshold for calling something an error", ">10x",
     "Below that, a mismatch is probably within the range. 01_ lidar at 50x "
     "clears it; 06_ cameras at 2x does not."),
]

# --- report section 7.2 / 7.3 ----------------------------------------------
# Driver C. Multiplier on SENSOR COUNTS after 2040. All scenarios are 1.00 at
# 2040 by construction, so Driver C cannot perturb any year for which sourced
# data exists (validation V9).
SCENARIO_ANCHORS = [
    # scenario, name, 2040, 2050, 2060, 2070, weight, mechanism
    ("S1", "Cost-driven convergence", 1.00, 1.00, 1.00, 1.00, 0.25,
     "Sensors get cheap, but sensing converges on a minimal efficient set "
     "(camera + radar). Redundancy achieved in compute, not hardware. The "
     "vision-only thesis."),
    ("S2", "Regulated redundancy", 1.00, 1.15, 1.28, 1.40, 0.50,
     "Cost decline makes redundancy affordable; L3/L4 type approval requires "
     "independent modality redundancy for fail-operational behaviour. MODE."),
    ("S3", "Full fail-operational", 1.00, 1.35, 1.70, 2.00, 0.25,
     "Private L4 requires dual-redundant sensing across every modality plus "
     "degraded-mode operation. Each sensing function duplicated."),
]
SCENARIO_YEARS = [2040, 2050, 2060, 2070]

# --- report section 8 -------------------------------------------------------
VALIDATION = [
    ("V1", "2025 length AB / CD / EF", "1392 / 2486 / 3546 m", "+/-1%", "existing anchor"),
    ("V2", "Metres per camera 2025 AB / CD / EF", "28 / 25 / 31 m", "+/-15%", "17_ / 06_"),
    ("V3", "EF ranging sensors (radar+lidar) 2025", "4.9", "+/-1.0", "report S2.2 measured"),
    ("V4", "Sales-weighted ranging sensors 2025", "3.0", "+/-1.0", "report S4.1 vs Yole 2.5"),
    ("V5", "Lidar-equipped share, Europe, 2025", "0.01", "+/-0.01", "report S2.4"),
    ("V6", "Tier shares sum to 1, every segment-year", "1.000", "1e-9", "construction"),
    ("V7", "Composed presence vs 01_ static labels 2025", "agree", "+/-0.15", "report S6.3"),
    ("V8", "ADAS share of total length 2025 AB/CD/EF", "6 / 9 / 11 %", "+/-3 pp", "interface doc S1"),
    ("V9", "Driver C multiplier at 2040, all scenarios", "1.000", "1e-9", "report S7.2"),
    ("V10", "Scenario weights sum to 1", "1.000", "1e-9", "construction"),
]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = HDR
        cell.border = THIN
        cell.alignment = WRAP


def widths(ws, *pairs):
    for col, w in pairs:
        ws.column_dimensions[col].width = w


def fill_for(tag):
    if tag.startswith("FACT"):
        return GREEN
    if tag in ("DERIVED", "DRIVER B"):
        return YELLOW
    return ORANGE


def build():
    wb = Workbook()

    # ---------------- Notes ----------------
    ws = wb.active
    ws.title = "Notes"
    ws["A1"] = "19_ADAS_sensor_adoption.xlsx"
    ws["A1"].font = TITLE
    rows = [
        ("", ""),
        ("Source", f"Every number here comes from {REPORT}. Section numbers in "
                   "each sheet point back to it. Change the report and this "
                   "file together, or they will diverge."),
        ("Regenerate", "python3 tools/make_19_adas_sensor_adoption.py"),
        ("", ""),
        ("Replaces", "18_ sheet Sensors_per_Level, which was keyed on SAE "
                     "CERTIFICATION level and entirely unsourced."),
        ("Why", "Wiring follows INSTALLED HARDWARE, not the certificate. A car "
                "does not grow a wire when a regulator grants liability "
                "transfer. Volvo's EX90 has 31 sensors and is certified L2; "
                "BMW's i7 had 25 and was certified L3."),
        ("", ""),
        ("Colours", "GREEN = fact, sourced and dated.  YELLOW = derived by a "
                    "stated rule.  ORANGE = assumption, no source. "
                    "Argue with orange first."),
        ("", ""),
        ("Three drivers", "A = hardware tier (sheet Tier_Shares). "
                          "B = lidar penetration (sheet Lidar), SEPARATE "
                          "because lidar tracks cost and Chinese competitive "
                          "pressure, not tier or certification. "
                          "C = post-2040 sensor-count scenario (sheet "
                          "Scenarios), which removes the need to forecast "
                          "unknown sensor modalities."),
        ("", ""),
        ("To switch scenario", "ONE cell: Scenarios!B5 (Active_Scenario). "
                               "SAMPLE = all three inside one run's band "
                               "(default, and what to quote). S1/S2/S3 = pin "
                               "one, outputs get a _S1/_S2/_S3 suffix. "
                               "No code editing."),
        ("", ""),
        ("On uncertainty", "See sheet Uncertainty. These are predictions, not "
                           "ground truth. Where two credible sources disagree, "
                           "that disagreement MEASURES the real uncertainty - "
                           "it is data about the band width, not an error to "
                           "be corrected. A mismatch inside the observed "
                           "spread is not evidence of a mistake."),
        ("", ""),
        ("Weakest input", "Driver C multipliers (report S9.2). The two "
                          "mechanisms behind them are observable; the specific "
                          "values 1.0 / 1.4 / 2.0 are not sourced. Largest "
                          "single lever on the 2070 answer."),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = BOLD
        c = ws.cell(row=i, column=2, value=v)
        c.alignment = WRAP
    widths(ws, ("A", 16), ("B", 110))

    # ---------------- Parameters ----------------
    ws = wb.create_sheet("Parameters")
    ws["A1"] = "Levers - report S4.2 and S5.3"
    ws["A1"].font = TITLE
    hdr = ["Parameter", "Value", "Meaning", "Tag"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=3, column=j, value=h)
    style_header(ws, 3, len(hdr))
    for i, (p, v, m, tag) in enumerate(PARAMS, start=4):
        ws.cell(row=i, column=1, value=p)
        c = ws.cell(row=i, column=2, value=v)
        c.fill = YELLOW
        c.border = THIN
        ws.cell(row=i, column=3, value=m).alignment = WRAP
        ws.cell(row=i, column=4, value=tag).fill = fill_for(tag)
    widths(ws, ("A", 30), ("B", 10), ("C", 78), ("D", 15))

    # ---------------- Tiers ----------------
    ws = wb.create_sheet("Tiers")
    ws["A1"] = "ADAS hardware tiers - report S3"
    ws["A1"].font = TITLE
    ws["A2"] = ("H3 and H4 are anchored on measured cars (EQS, i7, EX90). "
                "H0 is pinned by EU GSR-2. H1 and H2 are the weakest rows.")
    ws["A2"].alignment = WRAP
    hdr = ["Tier", "Name", "Cam_min", "Cam_max", "Radar_min", "Radar_max",
           "Ultra_min", "Ultra_max", "Lidar_min", "Lidar_max",
           "Ranging_mode", "SAE label seen", "Tag"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, len(hdr))
    for i, row in enumerate(TIER_DEF, start=5):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN
            if 3 <= j <= 11:
                c.fill = YELLOW
        ws.cell(row=i, column=13).fill = fill_for(row[-1])
    widths(ws, ("A", 7), ("B", 32), ("L", 16), ("M", 15))

    # ---------------- Tier_Shares ----------------
    ws = wb.create_sheet("Tier_Shares")
    ws["A1"] = "Driver A - tier share of new BEV sales (MODE) - report S4.1"
    ws["A1"].font = TITLE
    ws["A2"] = ("Mode only. The Min/Max band is a TIME SHIFT of "
                "Tier_Band_Shift_Y years on the whole curve (Parameters "
                "sheet) - this keeps shares summing to 1, which independent "
                "Min/Mode/Max columns would not. Rows sum to 1 by "
                "construction; sheet Validation V6 checks it.")
    ws["A2"].alignment = WRAP
    hdr = ["Segment", "Year"] + TIERS + ["Sum", "Ranging_derived"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, len(hdr))
    ranging = {t[0]: t[10] for t in TIER_DEF}
    r = 5
    for seg in SEGMENTS:
        for yr in YEARS:
            sh = TIER_SHARES[seg][yr]
            ws.cell(row=r, column=1, value=seg)
            ws.cell(row=r, column=2, value=yr)
            for j, v in enumerate(sh, start=3):
                c = ws.cell(row=r, column=j, value=v)
                c.fill = YELLOW
                c.border = THIN
                c.number_format = "0.00"
            col_first, col_last = get_column_letter(3), get_column_letter(7)
            c = ws.cell(row=r, column=8, value=f"=SUM({col_first}{r}:{col_last}{r})")
            c.number_format = "0.000"
            c.fill = GREY
            terms = "+".join(f"{get_column_letter(3 + k)}{r}*{ranging[t]}"
                             for k, t in enumerate(TIERS))
            c = ws.cell(row=r, column=9, value=f"={terms}")
            c.number_format = "0.00"
            c.fill = GREY
            r += 1
        r += 1
    widths(ws, ("A", 10), ("B", 8), ("H", 9), ("I", 16))

    # ---------------- Lidar ----------------
    ws = wb.create_sheet("Lidar")
    ws["A1"] = "Driver B - lidar-equipped share of new BEV sales, Europe - report S5.3"
    ws["A1"].font = TITLE
    ws["A2"] = ("SEPARATE from tier, because lidar tracks cost and Chinese "
                "competitive pressure - not certification and not segment. "
                "China reached 21% of NEVs in 2025 at ~$200/unit. Mode puts "
                "Europe there around 2032-33. The lag is SAMPLED "
                "(Parameters sheet), so China-leads-Europe-follows is a "
                "testable hypothesis, not a baked-in assumption. "
                "Min = 4D imaging radar substitutes; Max = cost collapse plus "
                "weather redundancy. Both cases are credible; the band holds both.")
    ws["A2"].alignment = WRAP
    hdr = ["Year", "Share_Min", "Share_Mode", "Share_Max", "Tag", "Basis"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, len(hdr))
    for i, (yr, lo, mo, hi, tag, basis) in enumerate(LIDAR, start=5):
        ws.cell(row=i, column=1, value=yr)
        for j, v in enumerate((lo, mo, hi), start=2):
            c = ws.cell(row=i, column=j, value=v)
            c.fill = YELLOW
            c.border = THIN
            c.number_format = "0.00"
        ws.cell(row=i, column=5, value=tag).fill = fill_for(tag)
        ws.cell(row=i, column=6, value=basis).alignment = WRAP
    widths(ws, ("A", 8), ("B", 11), ("C", 12), ("D", 11), ("E", 13), ("F", 52))

    # ---------------- Presence_per_Tier ----------------
    ws = wb.create_sheet("Presence_per_Tier")
    ws["A1"] = "presence(component | tier) - report S6.2"
    ws["A1"].font = TITLE
    ws["A2"] = ("Replaces the STATIC Std/Opt/Rare factor in 01_. Compose as:  "
                "presence(component, segment, year) = SUM_tier "
                "share(tier,segment,year) x presence(component|tier).  "
                "Read share() from Tier_Shares - do NOT recompute it, or the "
                "sensor and wiring models will silently diverge. "
                "Rows are the ADAS sheet of 01_VehicleElectronics.xlsx, in order.")
    ws["A2"].alignment = WRAP
    hdr = ["#", "Component"] + TIERS + ["Tag", "Basis"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, len(hdr))
    for i, row in enumerate(PRESENCE, start=5):
        name, h0, h1, h2, h3, h4, tag, basis = row
        ws.cell(row=i, column=1, value=i - 4)
        ws.cell(row=i, column=2, value=name)
        for j, v in enumerate((h0, h1, h2, h3, h4), start=3):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN
            c.fill = GREEN if v == LD else YELLOW
            if v != LD:
                c.number_format = "0.00"
        ws.cell(row=i, column=8, value=tag).fill = fill_for(tag)
        ws.cell(row=i, column=9, value=basis).alignment = WRAP
    # Footnotes go in a column clear of every data column. Putting them in a
    # data column makes them read back as extra records: a pandas filter on
    # "Component is not null" would return 13 rows for a 12-row table.
    r = len(PRESENCE) + 6
    ws.cell(row=r, column=NOTE_COL,
            value="WATCH: row 9 DECLINES as rows 10 and 11 rise - the smart "
                  "camera is absorbed into the domain controller. That is a "
                  "SUBSTITUTION. The static model cannot express it, and such "
                  "pairs must be drawn jointly, not independently.").alignment = WRAP
    ws.cell(row=r, column=NOTE_COL).font = BOLD
    widths(ws, ("A", 5), ("B", 36), ("H", 15), ("I", 42),
           (get_column_letter(NOTE_COL), 70))

    # ---------------- Scenarios: MOVED ----------------
    # Driver C now lives in Data/20_scenarios.xlsx, generated by
    # tools/make_20_scenarios.py. It is a PROJECT-WIDE switch and 19_ is
    # ADAS-specific, so keeping it here made 19_ the wrong source of truth for
    # the sensor and PCB models. 19_ keeps tiers, lidar and presence.

    # ---------------- Uncertainty ----------------
    ws = wb.create_sheet("Uncertainty")
    ws["A1"] = "How wide is the real uncertainty? Measured from source disagreement"
    ws["A1"].font = TITLE
    ws["A2"] = ("These are PREDICTIONS, not ground truth. Where two credible "
                "sources disagree about the same quantity, the gap between "
                "them is a measurement of the true uncertainty - it is data "
                "about how wide the band should be, not an error to correct. "
                "A model result that misses a reference by less than the "
                "spread below is NOT evidence of a bug.")
    ws["A2"].alignment = WRAP
    hdr = ["Quantity", "Source A", "Value A", "Source B", "Value B",
           "Spread", "What it tells us"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, len(hdr))
    for i, row in enumerate(UNCERTAINTY, start=5):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN
            if j == 7:
                c.alignment = WRAP
        ws.cell(row=i, column=6).font = BOLD

    r = len(UNCERTAINTY) + 7
    ws.cell(row=r, column=1, value="Calibration that follows").font = TITLE
    hdr2 = ["Kind of uncertainty", "Magnitude", "Consequence"]
    for j, h in enumerate(hdr2, start=1):
        ws.cell(row=r + 1, column=j, value=h)
    style_header(ws, r + 1, len(hdr2))
    for i, (k, mag, cons) in enumerate(UNCERTAINTY_LESSONS, start=r + 2):
        ws.cell(row=i, column=1, value=k)
        c = ws.cell(row=i, column=2, value=mag)
        c.font = BOLD
        c.fill = ORANGE
        ws.cell(row=i, column=3, value=cons).alignment = WRAP
    widths(ws, ("A", 34), ("B", 22), ("C", 20), ("D", 22), ("E", 16),
           ("F", 10), ("G", 62))

    # ---------------- Validation ----------------
    ws = wb.create_sheet("Validation")
    ws["A1"] = "Targets the code must reproduce - report S8"
    ws["A1"].font = TITLE
    ws["A2"] = ("Assert these in BevWiring.py so that a report drifting from "
                "the code becomes a TEST FAILURE, not a discovery six months "
                "later. Tolerances are set from the Uncertainty sheet - not "
                "tighter than the sources themselves agree.")
    ws["A2"].alignment = WRAP
    hdr = ["#", "Target", "Value", "Tolerance", "Source", "Last run", "Pass?"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, len(hdr))
    for i, row in enumerate(VALIDATION, start=5):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v).border = THIN
        ws.cell(row=i, column=6).fill = GREY
        ws.cell(row=i, column=7).fill = GREY
    widths(ws, ("A", 5), ("B", 42), ("C", 22), ("D", 12), ("E", 26),
           ("F", 12), ("G", 8))

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print("written:", build())
