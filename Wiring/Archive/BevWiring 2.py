"""
BEV Wiring Harness Monte Carlo Forecast, 2010-2070 -- v3
================================================================================
Builds on v2 (per-wire-category, 3-era, fully-triangular MC). This version
changes the OUTPUT layer to match the RAWCLICStockAndFlow / EVmodel
conventions, and adds sensitivity analysis:

  1. Every output row is tagged Segment + Drivetrain='BEV' (this dataset is
     BEV-only; no placeholder rows for HEV/PHEV/Diesel/Petrol).
  2. Stats per (Segment, Wire_Type, Metric, Year): Mean, P2_5, Median, Mode,
     P97_5. Mode = the 50-bin histogram bin with the highest count (i.e. the
     mode is read off the SAME 50-bin histogram that's exported, not a
     separate KDE estimate -- self-consistent).
  3. 50-bin histograms exported as one tidy CSV (Segment, Wire_Type, Metric,
     Year, Bin_Left, Bin_Right, Count), at a set of snapshot years.
  4. Per-category plots are SEPARATED (small-multiples grid, one axes per
     wire type), not stacked/combined -- one grid figure per (segment, metric).
  5. Sensitivity / tornado plots (Spearman rank correlation, signed, sorted
     by |rho|), two views:
       a) parameter tornado -- for a given (wire_type, segment, metric, year),
          which of the 4 stochastic MC inputs (trim percentile, start
          fraction, target fraction, decay tau) drives ITS OWN uncertainty.
       b) contribution tornado -- for a given (segment, metric, year), which
          of the 28 wire categories correlates most with the SEGMENT TOTAL
          (i.e. which categories drive total uncertainty).

Core trajectory model (era A/B/C, classification, blending) is UNCHANGED
from v2. The rationale used to live only in v2's docstring; it is now
documented in full below and inline at each constant, so this file stands
on its own.


HOW THE MODEL WORKS (read this first if you are coming back to this file)
================================================================================
Input data gives, for each (wire type x vehicle segment), a min/mode/max
triple at THREE anchor years only: 2020, 2025, 2030. Everything from 2010
to 2070 is interpolated or extrapolated from those three anchors. The
timeline is split into three eras, then smoothly blended:

    era A: 2010 -> 2020   backcast. Linear-ish ramp (smoothstep) from
                          `START_FRACTION_TRI` x (2020 value) up to the
                          2020 anchor.
    era B: 2020 -> 2030   interpolation between the three anchors with a
                          shape-preserving cubic Hermite spline
                          (`hermite3_eval`) -- no overshoot between knots.
    era C: 2030 -> 2070   forecast. Exponential relaxation from the 2030
                          anchor toward a long-run asymptote
                          `CLASS_TARGET_TRI` x (2030 value), with e-folding
                          time `CLASS_TAU_TRI`.

Eras are cross-faded, not hard-switched: `BLEND_HALFWIDTH_AB` / `_BC` set
the width of the smoothstep cross-fade at the 2020 and 2030 seams, so the
trajectory has no kink there.

Era C's shape is chosen per wire type by `classify()`, which compares the
2030 mode to the 2020 mode and labels each wire type declining / stable /
growing / emerging / absent. That label selects which `CLASS_*_TRI` row is
used. So: the DATA decides the class, the class decides the future.

FOUR stochastic inputs are drawn per Monte Carlo iteration. These are
exactly the four that the parameter tornado plots rank:

    1. trim_percentile   (u)  -- where in the min/mode/max triangle this
                                 vehicle sits. Drawn ONCE PER SEGMENT and
                                 shared across all 28 wire types (see
                                 TRIM_JITTER_STD -- this is the single most
                                 consequential assumption in the file).
    2. start_fraction         -- era A backcast depth.
    3. target_fraction        -- era C asymptote.
    4. decay_tau              -- era C speed.

A fifth, `HEIGHT_ADDER_TRI`, is drawn per segment to derive the J* segments.

UNITS: the two metrics are named "Length (m)" and "Cu (kg)" -- quantity with
its unit in parentheses, which is the naming convention for this project.
That exact string is what appears in the Metric column of both output CSVs
and on plot titles/axes, so a chart is readable without consulting this file.
Both are PER VEHICLE, not fleet totals.

WHERE TO CHANGE THINGS: every tunable lives in section 0 below and is
annotated with its unit, why it has the value it has, and what moves if you
change it. A few magic numbers are unavoidably hardcoded inside the era-B
spline (`hermite3_eval`) because they encode the 2020/2025/2030 anchor
spacing itself -- those are flagged in that function's docstring.
"""

from __future__ import annotations

import numpy as np
import pandas as pd
from dataclasses import dataclass, field
from pathlib import Path
from scipy.stats import spearmanr
from openpyxl import load_workbook

# --------------------------------------------------------------------------
# 0. CONFIGURATION
#
#    Every constant below is annotated as:
#        WHAT it is | UNIT | WHY this value | WHAT MOVES if you change it
#    "tri" in a name always means a triangular distribution given as the
#    3-tuple (min, mode, max) -- i.e. (lower bound, most likely, upper
#    bound), fed to numpy's rng.triangular in that exact order.
# --------------------------------------------------------------------------

# Folder layout (common project folder, sibling subfolders):
#   <common>/
#     Wiring/   <- this script lives here
#     Data/     <- 15_BEV_wiring_harness_data.xlsx lives here
SCRIPT_DIR = Path(__file__).resolve().parent
COMMON_DIR = SCRIPT_DIR.parent
DATA_DIR = COMMON_DIR / "Data"
# NOTE: the "15_" prefix is part of the actual filename in Data/ (the repo
# numbers its data files). Without it this script dies with FileNotFoundError.
DATA_FILE = DATA_DIR / "15_BEV_wiring_harness_data.xlsx"
WIRE_LEVEL_SHEET = "1_Wire-Level Data"   # tab name from the split-to-Excel workbook
# ^ that sheet holds 28 wire types x 3 segments x 3 time periods = 252 rows.
#   Its header is NOT on row 1 (there are title/subtitle/blank rows above),
#   which is why _find_header_row() hunts for the "Wire_Type" cell.

OUT_DIR = SCRIPT_DIR / "outputs"
OUT_SUBDIRS = {
    "data":        OUT_DIR / "data",                     # stats + histogram CSVs
    "totals":      OUT_DIR / "plots" / "totals",         # overall summary (all segments)
    "trajectories": OUT_DIR / "plots" / "trajectories",  # per-wire-type grids (28)
    "histograms":  OUT_DIR / "plots" / "histograms",     # per-wire-type histogram grids
    "sensitivity": OUT_DIR / "plots" / "sensitivity",    # tornado plots (both kinds)
    "category_trajectories": OUT_DIR / "plots" / "category_trajectories",  # 7 wire-category groups
    "category_histograms":   OUT_DIR / "plots" / "category_histograms",
    "category_overview":     OUT_DIR / "plots" / "category_overview",      # groups vs. total, per segment
}
for _p in OUT_SUBDIRS.values():
    _p.mkdir(parents=True, exist_ok=True)

# --- Run extent -------------------------------------------------------------
START_YEAR = 2010      # first output year | UNIT: calendar year (CE).
                       # 2010 is a BACKCAST: there is no 2010 data. Everything
                       # before 2020 is reconstructed via START_FRACTION_TRI.
                       # Treat pre-2020 output as indicative, not evidenced.
END_YEAR = 2070        # last output year | UNIT: calendar year (CE).
                       # 40 years past the last anchor (2030). The further past
                       # 2030, the more the answer is just CLASS_TARGET_TRI.
N_ITER = 3000          # Monte Carlo iterations | UNIT: count (dimensionless).
                       # Drives all percentile stability. 3000 gives ~+/-1%
                       # noise on P2.5/P97.5. Runtime and memory scale linearly
                       # (every trajectory array is N_ITER x 61 float64).
RANDOM_SEED = 42       # RNG seed | UNIT: none (integer seed).
                       # Fixed so runs are reproducible. CHANGE THIS to check
                       # that your conclusions are not a single-seed artefact.

# --- Vehicle segments -------------------------------------------------------
# European ("Euro Car Segment") size classes. Per the data workbook's
# methodology sheet these correspond to overall vehicle length / wheelbase:
#     AB -> 3.7-4.2 m overall, 2.5 m wheelbase   (city / small)
#     CD -> 4.3-4.8 m overall, 2.7 m wheelbase   (compact / mid)
#     EF -> >4.8 m overall,    3.1 m wheelbase   (large / luxury)
# Longer vehicle => longer harness runs, hence segment matters at all.
BASE_SEGMENTS = ["AB", "CD", "EF"]   # the ONLY segments present in the input data

# Derived (not measured) segments: the J-segment / taller-body counterpart of
# each base segment. These have NO input data of their own -- they are produced
# purely by scaling the base segment by (1 + HEIGHT_ADDER_TRI). See that
# constant for the reasoning.
#   ASSUMPTION TO CONFIRM: "J" is read here as the SUV/taller-body class in the
#   Euro segment scheme. If that is not your intent, this mapping is the only
#   place it is encoded.
J_SUFFIX = {"AB": "JA-JB", "CD": "JC-JD", "EF": "JE-JF"}

# The two quantities modelled, per vehicle. Both are PER-VEHICLE, not fleet.
# NAMING CONVENTION: "Quantity (unit)" -- the unit travels with the name, so
# it is visible in the CSV Metric column and on every plot axis and title.
#   "Length (m)" | metres of wire per vehicle
#   "Cu (kg)"    | kilograms of copper CONDUCTOR per vehicle
#                (conductor only -- excludes insulation, connectors, shields'
#                non-Cu content. Workbook derives it as AWG cross-section x
#                8.96 g/cm3 Cu density x ~0.88-0.92 stranding fill factor.)
# These strings are also the keys used throughout raw_total / raw_cat dicts,
# and they are mapped to workbook column names in build_category_anchor_table.
METRICS = ["Length (m)", "Cu (kg)"]

# Filename-safe slug for each metric. Output PNGs are named from this, not
# from METRICS directly, because the display names contain spaces and
# parentheses which make for awkward paths and shell-quoting.
METRIC_SLUG = {"Length (m)": "length_m", "Cu (kg)": "cu_kg"}

DRIVETRAIN = "BEV"     # constant tag stamped on every output row | UNIT: label.
                       # This dataset is BEV-only; the column exists so outputs
                       # concatenate cleanly with the RAWCLICStockAndFlow /
                       # EVmodel tables, which do carry other drivetrains.

# --- Anchor years -----------------------------------------------------------
# The workbook labels its three time periods as ranges; the model needs a
# single year per period. Each range is collapsed to ONE representative year:
PERIOD_YEAR = {
    "Historical_2020-2023": 2020,   # range collapsed to its start
    "Current_2025":         2025,
    "Future_2026-2030":     2030,   # range collapsed to its end
}
# Same three years as a float array, used to clamp era B to the interpolation
# range (outside 2020-2030 the spline must not be evaluated -- eras A and C
# take over there). UNIT: calendar years (CE), must stay sorted ascending.
# WARNING: these three values are ALSO hardcoded inside hermite3_eval (the
# 5-year spacing and the 2020/2025 knots). Changing the anchor years here
# alone is NOT enough -- see that function's docstring.
ANCHOR_YEARS = np.array([2020.0, 2025.0, 2030.0])

# --- Era A: the 2010 -> 2020 backcast ---------------------------------------
# Value at START_YEAR expressed as a FRACTION OF THAT WIRE TYPE'S 2020 VALUE.
# UNIT: dimensionless ratio (0.70 => a 2010 vehicle had 70% of its 2020 wiring).
# WHY: harnesses grew steadily through the 2010s with electrification and
# content creep, so 2010 < 2020, but not dramatically. The wide 0.50-0.90 band
# is honest about this being unevidenced.
# EFFECT: raises/lowers the whole left-hand shoulder of every curve. Applies
# uniformly to ALL wire types -- it is not class-dependent, unlike era C.
START_FRACTION_TRI = (0.50, 0.70, 0.90)

# --- Era C: the 2030 -> 2070 forecast ---------------------------------------
# THE LONG-RUN ASYMPTOTE, as a MULTIPLE OF THAT WIRE TYPE'S 2030 VALUE.
# UNIT: dimensionless ratio. 1.0 = flat at the 2030 level; 0.55 = settles at
# 55% of 2030; 8.0 = grows to 8x its 2030 level.
# This is the single biggest lever on any post-2040 number in this model.
# The class is assigned by classify() from the data, then this table decides
# where that class ends up:
CLASS_TARGET_TRI = {
    # wire types being designed out (zonal architecture removes point-to-point
    # runs, wireless BMS removes sense wires, CAN/LIN give way to Ethernet).
    # Decline continues past 2030 but does NOT go to zero -- floor ~30%.
    "declining": (0.30, 0.55, 0.80),
    # mature, physically-driven wiring (mostly HV power routing, which the
    # workbook notes is architecture-independent). Roughly flat; the 0.85-1.35
    # band allows mild shrink or mild growth.
    "stable":    (0.85, 1.05, 1.35),
    # already present in 2020 and growing strongly (e.g. Ethernet backbone).
    # Wide 1.5-6.0 band because saturation timing is genuinely unknown.
    "growing":   (1.5, 3.0, 6.0),
    # ~absent in 2020, appearing by 2030. Multiplies a SMALL 2030 base, so a
    # big multiplier is not as wild as it looks -- but it is the widest and
    # least evidenced band in the file. Expect these to dominate the parameter
    # tornado plots.
    "emerging":  (3.0, 8.0, 20.0),
    # never present in this segment (e.g. rear-motor cabling on a single-motor
    # car). Forced to exactly zero; these rows are dropped from all outputs.
    "absent":    (0.0, 0.0, 0.0),
}

# HOW FAST era C approaches the asymptote above. e-folding time of the
# exponential relaxation: value(t) = target + (v2030 - target) * exp(-(t-2030)/tau).
# UNIT: YEARS. tau=25 => ~63% of the way to the target after 25 years,
# ~86% after 50. Bigger tau = slower, more gradual change.
CLASS_TAU_TRI = {
    "declining": (15, 25, 40),   # designing wiring out takes platform cycles
    "stable":    (15, 25, 40),   # slow drift either way
    "growing":   (8, 15, 25),    # faster -- tech adoption S-curves are steeper
    "emerging":  (8, 15, 25),    # faster, same reason
    "absent":    (1, 1, 1),      # irrelevant (target and value are both 0);
                                 # 1 not 0 purely to avoid a divide-by-zero.
}

# --- classify(): thresholds that assign each wire type to a class above -----
# Both compare the 2030 MODE to the 2020 MODE: ratio = mode_2030 / mode_2020.
# UNIT: dimensionless ratio.
GROW_RATIO_THRESH = 1.4      # ratio >= 1.4 (i.e. +40% or more) -> "growing"
DECLINE_RATIO_THRESH = 0.7   # ratio <= 0.7 (i.e. -30% or more) -> "declining"
                             # anything between 0.7 and 1.4        -> "stable"
# Below this, a value counts as "zero" rather than a small real number.
# UNIT: same as the metric being tested (metres or kg) -- it is an ABSOLUTE
# floor, not relative. Guards the mode_2030/mode_2020 division and separates
# "emerging" (zero in 2020, nonzero in 2030) from "absent" (zero in both).
ZERO_EPS = 1e-6

# --- J* derived segments ----------------------------------------------------
# Uplift applied to a base segment to synthesise its J* counterpart:
#     J_value = base_value * (1 + height_adder)
# UNIT: dimensionless fraction (0.10 => +10% wire length and +10% copper).
# WHY: a taller/larger body on the same platform means longer vertical runs
# (roof, tailgate, raised floor) but the same electronic content.
# NOTE: drawn ONCE PER SEGMENT PER ITERATION, then applied to both metrics and
# to every wire type in that segment -- so J* is a perfectly correlated,
# rigidly scaled copy of its base segment, never an independent estimate.
HEIGHT_ADDER_TRI = (0.05, 0.10, 0.15)

# --- Era seam blending ------------------------------------------------------
# Half-width of the smoothstep cross-fade at each era boundary.
# UNIT: YEARS. 3 => era A and era B are mixed across 2017-2023 (2020 +/- 3),
# reaching 50/50 exactly at 2020. Likewise 2027-2033 for the B/C seam.
# WHY: prevents a visible kink/discontinuity where the eras meet.
# EFFECT: larger = smoother but the anchors are honoured less exactly;
# 0 would mean a hard switch at the seam.
BLEND_HALFWIDTH_AB = 3   # blends era A (backcast) into era B (spline) at 2020
BLEND_HALFWIDTH_BC = 3   # blends era B (spline) into era C (forecast) at 2030

# --- Cross-wire-type correlation (IMPORTANT -- read this one) ---------------
# The "trim percentile" u decides where in the min/mode/max triangle a given
# simulated vehicle sits: u=0.1 is a sparsely-wired example of that segment,
# u=0.9 a heavily-optioned one.
#
# u is drawn ONCE PER SEGMENT and REUSED for all 28 wire types (see the
# `u_trim` dict in run_monte_carlo). That is a deliberate modelling choice: it
# says a heavily-optioned car is heavily wired EVERYWHERE, rather than letting
# 28 independent draws average each other out. Independent draws would shrink
# the segment-total uncertainty band by roughly sqrt(28) and would be wrong.
#
# TRIM_JITTER_STD then adds a small per-wire-type Gaussian wobble on top:
#     u_wire = clip(u_segment + Normal(0, TRIM_JITTER_STD), 0, 1)
# UNIT: standard deviation in PERCENTILE units, i.e. same units as u itself,
# on a 0-1 scale (0.05 = 5 percentile points).
# EFFECT: 0.0 = perfect rank correlation across wire types (widest total band);
# large = wire types decouple and the total band narrows. 0.05 keeps them
# strongly but not perfectly coupled.
TRIM_JITTER_STD = 0.05

# --- Output/reporting knobs (affect what is EXPORTED, not the model) --------
# Years at which the full 50-bin distribution is written out. The stats CSV
# covers every year 2010-2070; the histogram CSV would be enormous at that
# resolution, so it is restricted to these snapshots.
# UNIT: calendar years (CE). Must lie within START_YEAR..END_YEAR.
SNAPSHOT_YEARS = [2010, 2020, 2025, 2030, 2040, 2050, 2060, 2070]

# Histogram resolution | UNIT: count of bins (dimensionless).
# Used for BOTH the exported histograms AND the reported Mode -- the Mode is
# the centre of the fullest of these bins, so changing this changes the Mode
# column. That coupling is intentional (the Mode is reproducible from the
# exported histogram), but it does mean Mode is quantised to bin width.
N_HIST_BINS = 50

# Years for the sensitivity/tornado plots | UNIT: calendar years (CE).
# 2030 = last anchored year, 2050/2070 = increasingly assumption-driven.
TORNADO_YEARS = [2030, 2050, 2070]

# The uncertainty display range: (lower, upper) percentile of the Monte
# Carlo distribution reported as the band. UNIT: percent (0-100), not a
# 0-1 fraction. [2.5, 97.5] = central 95%; [5, 95] = central 90%.
# NOTE: this line is NOT read by the code. build_stats_table() and the
# plotting functions each hardcode [2.5, 50, 97.5] inline, so changing it
# alone has no effect -- the inline calls must be edited too.
STAT_PCTS = [2.5, 97.5]   # P2_5, P97_5 -- Median/Mean/Mode computed separately

# !! UNUSED -- kept only as documentation of intent. !!
# The authoritative copy of these four names is the `params` dict inside
# param_tornado_data(); this list is not read. It documents the four
# stochastic inputs whose influence the parameter tornado ranks.
PARAM_NAMES = ["trim_percentile", "start_fraction", "target_fraction", "decay_tau"]


# --------------------------------------------------------------------------
# 1. LOAD & CLASSIFY  (v4: reads from the Excel workbook, not the raw CSV)
# --------------------------------------------------------------------------

def _find_header_row(ws, first_col_name: str) -> int:
    """Locate the 1-indexed row whose column-A value matches `first_col_name`.
    Robust to the sheet's title/note/blank rows above the actual header,
    so it survives minor layout changes in the workbook."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == first_col_name:
            return r
    raise ValueError(f"Could not find header row starting with '{first_col_name}' "
                      f"on sheet '{ws.title}'.")


def load_wire_level_data(xlsx_path: Path = DATA_FILE,
                          sheet_name: str = WIRE_LEVEL_SHEET) -> pd.DataFrame:
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"Data file not found: {xlsx_path}\n"
            f"Expected layout: <common>/Wiring/<this script>, <common>/Data/{xlsx_path.name}"
        )
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {xlsx_path}. "
                          f"Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]
    header_row = _find_header_row(ws, "Wire_Type")
    wb.close()

    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=header_row - 1)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    df = df[df["Wire_Type"].notna()]  # drop any stray blank rows below the table
    df["Year"] = df["Time_Period"].map(PERIOD_YEAR)
    return df


def classify(hist_mode: float, fut_mode: float) -> str:
    """Label a wire type by comparing its 2030 mode to its 2020 mode.

    The returned label selects the row used from CLASS_TARGET_TRI and
    CLASS_TAU_TRI, i.e. it decides the entire post-2030 shape. Order of the
    checks matters: the two zero cases are resolved BEFORE the ratio is
    computed, so the division can never be by zero.

    Args:
        hist_mode: mode value at the 2020 anchor. UNIT: metres or kg.
        fut_mode:  mode value at the 2030 anchor. UNIT: same as hist_mode.

    Returns:
        'absent'    -- zero at both anchors; never fitted to this segment.
        'emerging'  -- zero in 2020, present in 2030 (ratio is undefined).
        'growing'   -- ratio >= GROW_RATIO_THRESH.
        'declining' -- ratio <= DECLINE_RATIO_THRESH.
        'stable'    -- anything in between.
    """
    if fut_mode < ZERO_EPS and hist_mode < ZERO_EPS:
        return "absent"
    if hist_mode < ZERO_EPS:
        return "emerging"
    ratio = fut_mode / hist_mode
    if ratio >= GROW_RATIO_THRESH:
        return "growing"
    if ratio <= DECLINE_RATIO_THRESH:
        return "declining"
    return "stable"


@dataclass
class CategoryAnchors:
    """The model's input for ONE (wire type, segment, metric): a min/mode/max
    triple at each of the three anchor years, plus its class label.

    There is one of these per combination, so 28 wire types x 3 segments x
    2 metrics = 168 instances (minus none -- 'absent' ones are kept here and
    filtered out at output time).
    """
    wire_type: str       # e.g. "HV_Battery_Main" -- the finest granularity (28 of them)
    wire_category: str   # e.g. "HV Power Cable" -- the group it rolls up into (7 of them)
    segment: str         # "AB" | "CD" | "EF" (base segments only; J* is derived later)
    metric: str          # "Length (m)" | "Cu (kg)" -- also fixes the UNIT of lo/mode/hi
    lo: np.ndarray       # (3,) minimum at [2020, 2025, 2030]. UNIT: m or kg per metric
    mode: np.ndarray     # (3,) most likely, same order/unit
    hi: np.ndarray       # (3,) maximum,     same order/unit
    cls: str             # class from classify(); selects the CLASS_*_TRI row used


def build_category_anchor_table(df: pd.DataFrame) -> list[CategoryAnchors]:
    # metric display name -> the (min, mode, max) column names in the workbook
    col = {"Length (m)": ("Length_Min_m", "Length_Mode_m", "Length_Max_m"),
           "Cu (kg)":    ("Cu_Total_Min_kg", "Cu_Total_Mode_kg", "Cu_Total_Max_kg")}
    out = []
    for (wt, seg), g in df.groupby(["Wire_Type", "Vehicle_Segment"]):
        g = g.set_index("Year").sort_index()
        wire_category = g["Wire_Category"].iloc[0]
        for metric, (c_lo, c_mode, c_hi) in col.items():
            lo = g.loc[[2020, 2025, 2030], c_lo].to_numpy(dtype=float)
            mode = g.loc[[2020, 2025, 2030], c_mode].to_numpy(dtype=float)
            hi = g.loc[[2020, 2025, 2030], c_hi].to_numpy(dtype=float)
            cls = classify(mode[0], mode[2])
            out.append(CategoryAnchors(wt, wire_category, seg, metric, lo, mode, hi, cls))
    return out


# --------------------------------------------------------------------------
# 2. SAMPLING / SPLINE / BLENDING  (identical to v2)
# --------------------------------------------------------------------------

def triangular_ppf(u, lo, mode, hi):
    """Inverse CDF (quantile function) of a triangular distribution.

    Maps a percentile u in [0,1] to the value at that percentile of
    Triangular(lo, mode, hi). This is NOT the same as drawing a random
    triangular sample: the caller supplies u, which is what lets the same
    percentile be reused across the three anchor years and across wire types.

    WHY THIS EXISTS (rather than rng.triangular): build_category_trajectories
    calls it three times -- once per anchor year -- with the SAME u. That makes
    the 2020/2025/2030 draws perfectly rank-correlated, so a simulated vehicle
    that is in the 80th percentile in 2020 is also in the 80th percentile in
    2025 and 2030. Independent draws per anchor would produce implausible
    zig-zag trajectories that jump between the low and high edges of the band.

    Args:
        u:    percentile(s) in [0,1]. UNIT: dimensionless.
        lo, mode, hi: triangle parameters. UNIT: whatever the metric is
              (metres or kg). Broadcast against u.

    Returns:
        Value at percentile u. Degenerate triangles (hi == lo, i.e. a fixed
        point, which is how 'absent' rows arrive here) return lo exactly
        rather than dividing by a zero span.
    """
    u = np.asarray(u, dtype=float); lo = np.asarray(lo, dtype=float)
    mode = np.asarray(mode, dtype=float); hi = np.asarray(hi, dtype=float)
    span = hi - lo
    span_safe = np.where(span <= 0, 1.0, span)
    fc = (mode - lo) / span_safe
    left = u < fc
    out = np.where(
        left,
        lo + np.sqrt(np.clip(u * span_safe * (mode - lo), 0, None)),
        hi - np.sqrt(np.clip((1 - u) * span_safe * (hi - mode), 0, None)),
    )
    return np.where(span <= 0, lo, out)


def sample_triangular_tri(rng, tri, size):
    """Draw `size` independent samples from a (min, mode, max) 3-tuple.

    Convenience wrapper that unpacks the *_TRI constants in the order numpy
    expects. Unlike triangular_ppf, these draws are INDEPENDENT -- used for the
    parameters that should vary freely per iteration (start_fraction,
    target_fraction, tau, height_adder), not for the anchor values.
    """
    lo, mode, hi = tri
    return rng.triangular(lo, mode, hi, size=size)


def hermite3_eval(t, y0, y1, y2):
    """Era B: shape-preserving cubic Hermite interpolation through the three
    anchors (2020, 2025, 2030), evaluated for all MC iterations at once.

    Given three anchor values per iteration, builds two cubic segments
    (2020-2025 and 2025-2030) that pass exactly through all three points.

    WHY HERMITE AND NOT A PLAIN CUBIC/SPLINE: the interior slope d1 uses the
    HARMONIC mean of the two neighbouring secant slopes (the Fritsch-Carlson
    monotonicity rule), and is forced to ZERO when the two secants have
    opposite signs -- i.e. at a local peak or trough. That guarantees the curve
    never overshoots the anchor values. A natural cubic spline would happily
    swing below zero between two small anchors, producing negative wire length.

    NOTE ON THE ANCHOR YEARS BELOW (documentation only -- nothing to change
    unless YOU decide to):
    A spline needs to know not just the anchor VALUES but WHERE the anchors
    sit on the time axis -- to compute 2023 it must know 2023 is 60% of the
    way from 2020 to 2025. That time-axis information is written directly
    into the code below as three literals:
        h = 5.0     the spacing between anchors, in years
        2020.0      the first knot
        2025.0      the middle knot
    They are NOT read from ANCHOR_YEARS. ANCHOR_YEARS is used only by the
    CALLER, to clip `t` to the interpolation range. So if the anchor years
    are ever changed via PERIOD_YEAR / ANCHOR_YEARS, these three literals
    stay behind at the old years and the curve is laid out on the wrong time
    axis -- with no error raised. Flagged here so the coupling is visible.

    Args:
        t:  years to evaluate. UNIT: calendar years (CE), shape (T,).
            Callers pass values already clipped to 2020..2030.
        y0, y1, y2: anchor values at 2020 / 2025 / 2030, one per MC iteration.
            UNIT: metres or kg, each shape (N_ITER,).

    Returns:
        Array (N_ITER, T) of interpolated values. UNIT: same as y0/y1/y2.
    """
    h = 5.0
    # Secant slopes across each interval. UNIT: metres/year or kg/year.
    delta0 = (y1 - y0) / h
    delta1 = (y2 - y1) / h
    # Slope at the MIDDLE knot: zero if the two secants disagree in sign (a
    # peak or trough), else their harmonic mean -- the Fritsch-Carlson rule
    # that keeps the curve monotone and stops it overshooting the anchors.
    same_sign = (delta0 * delta1) > 0
    with np.errstate(divide="ignore", invalid="ignore"):
        harmonic = 2.0 / (1.0 / np.where(delta0 == 0, np.nan, delta0)
                           + 1.0 / np.where(delta1 == 0, np.nan, delta1))
    d1 = np.nan_to_num(np.where(same_sign, harmonic, 0.0), nan=0.0)
    # Slopes at the two OUTER knots: one-sided differences (no neighbour on
    # the far side to average with).
    d0 = delta0
    d2 = delta1
    N = y0.shape[0]; T = t.shape[0]
    tt = t[None, :]

    def hermite_seg(s, ya, yb, da, db):
        """One cubic segment. `s` is position within the segment on a 0-1
        scale; h converts the per-year slopes da/db into per-unit-s slopes."""
        h00 = 2 * s**3 - 3 * s**2 + 1
        h10 = s**3 - 2 * s**2 + s
        h01 = -2 * s**3 + 3 * s**2
        h11 = s**3 - s**2
        return (h00 * ya[:, None] + h10 * h * da[:, None]
                + h01 * yb[:, None] + h11 * h * db[:, None])

    seg1 = tt <= 2025.0                          # left of the middle knot?
    s1 = np.clip((tt - 2020.0) / h, 0.0, 1.0)
    val1 = hermite_seg(s1, y0, y1, d0, d1)
    s2 = np.clip((tt - 2025.0) / h, 0.0, 1.0)
    val2 = hermite_seg(s2, y1, y2, d1, d2)
    return np.where(np.broadcast_to(seg1, (N, T)), val1, val2)


def smoothstep(t, edge0, edge1):
    """Classic 3x^2-2x^3 smoothstep: 0 below edge0, 1 above edge1, with a
    smooth S-shaped ramp between (zero first derivative at both ends).

    Used for two different jobs in this model:
      1. the era A ramp from the 2010 start value up to the 2020 anchor;
      2. the cross-fade weights at the era seams (BLEND_HALFWIDTH_AB / _BC),
         where it returns the WEIGHT given to the later era.

    Args:
        t:      years to evaluate. UNIT: calendar years (CE).
        edge0:  year at which the ramp starts (output 0).
        edge1:  year at which it completes (output 1). Must be > edge0.

    Returns:
        Weights in [0,1]. UNIT: dimensionless.
    """
    x = np.clip((t - edge0) / (edge1 - edge0), 0.0, 1.0)
    return x * x * (3 - 2 * x)


# --------------------------------------------------------------------------
# 3. PER-CATEGORY TRAJECTORY -- now also returns the sampled input arrays,
#    needed for sensitivity analysis.
# --------------------------------------------------------------------------

@dataclass
class CategoryDraw:
    """One wire type x segment x metric, simulated N_ITER times.

    Carries both the resulting trajectories AND the four stochastic inputs
    that produced them -- the inputs are retained solely so the parameter
    tornado (param_tornado_data) can rank their influence afterwards.
    """
    traj: np.ndarray            # the trajectories. (N_ITER, n_years). UNIT: m or kg
    u_eff: np.ndarray           # (N_ITER,) trim percentile ACTUALLY used, i.e. after
                                # the TRIM_JITTER_STD wobble and clipping to [0,1].
                                # UNIT: dimensionless percentile in [0,1].
    start_fraction: np.ndarray  # (N_ITER,) era A backcast depth, as a fraction of the
                                # 2020 value. UNIT: dimensionless ratio.
    target_fraction: np.ndarray # (N_ITER,) era C asymptote, as a multiple of the
                                # 2030 value. UNIT: dimensionless ratio.
    tau: np.ndarray             # (N_ITER,) era C e-folding time. UNIT: years.


def build_category_trajectories(cat, years, rng, u_trim, n_iter) -> CategoryDraw:
    """Simulate N_ITER trajectories for ONE (wire type, segment, metric).

    Assembles the three eras described in the module docstring and cross-fades
    them. All four stochastic inputs are drawn here.

    Args:
        cat:    CategoryAnchors -- the min/mode/max at each of the 3 anchor
                years, plus the class label from classify().
        years:  every output year. UNIT: calendar years (CE), shape (n_years,).
        rng:    the shared numpy Generator. Consuming draws from it in a
                different order changes results even at a fixed RANDOM_SEED.
        u_trim: (n_iter,) the SEGMENT-level trim percentile, shared with every
                other wire type in this segment -- see TRIM_JITTER_STD.
        n_iter: number of MC iterations.

    Returns:
        CategoryDraw with trajectories clipped at >= 0.
    """
    n_years = len(years)
    # 'absent' = this wire type does not exist in this segment. Short-circuit to
    # all-zeros: no RNG is consumed, and the row is later dropped from outputs.
    if cat.cls == "absent":
        z = np.zeros(n_iter)
        return CategoryDraw(np.zeros((n_iter, n_years)), z, z, z, z)

    # (1) Trim percentile: segment-level u + small per-wire-type wobble.
    u = np.clip(u_trim + rng.normal(0, TRIM_JITTER_STD, size=n_iter), 0.0, 1.0)

    # Same u at all three anchors => rank-correlated across time (see
    # triangular_ppf). y0/y1/y2 are this iteration's 2020/2025/2030 values.
    y0 = triangular_ppf(u, cat.lo[0], cat.mode[0], cat.hi[0])
    y1 = triangular_ppf(u, cat.lo[1], cat.mode[1], cat.hi[1])
    y2 = triangular_ppf(u, cat.lo[2], cat.mode[2], cat.hi[2])

    # --- era B (2020-2030): spline through the anchors. Clipping t means the
    #     spline is flat-extrapolated outside 2020-2030; that is harmless
    #     because the blend weights below hand over to eras A and C there.
    t_clip = np.clip(years, ANCHOR_YEARS[0], ANCHOR_YEARS[-1])
    f_B = hermite3_eval(t_clip, y0, y1, y2)

    # --- era A (2010-2020): (2) start_fraction. Smoothstep ramp from
    #     start_fraction * (2020 value) up to the 2020 value.
    start_fraction = sample_triangular_tri(rng, START_FRACTION_TRI, n_iter)
    start_val = start_fraction * y0
    s_a = smoothstep(years, START_YEAR, 2020.0)[None, :]
    f_A = start_val[:, None] + (y0[:, None] - start_val[:, None]) * s_a

    # --- era C (2030-2070): (3) target_fraction and (4) tau. Exponential
    #     relaxation from the 2030 value toward target_fraction * (2030 value).
    #     NOTE: for years < 2030 dt is negative, so exp(-dt/tau) > 1 and f_C
    #     diverges backwards in time. That is never visible because w_bc below
    #     is 0 until 2030-BLEND_HALFWIDTH_BC. If you ever widen that blend a
    #     long way, this term is what will misbehave first.
    target_fraction = sample_triangular_tri(rng, CLASS_TARGET_TRI[cat.cls], n_iter)
    target_val = target_fraction * y2
    tau = sample_triangular_tri(rng, CLASS_TAU_TRI[cat.cls], n_iter)
    dt = (years - 2030.0)[None, :]
    f_C = target_val[:, None] + (y2[:, None] - target_val[:, None]) * np.exp(-dt / tau[:, None])

    # --- cross-fade the seams: A into B at 2020, then that into C at 2030.
    #     w is the weight of the LATER era, 0.5 exactly at the seam year.
    w_ab = smoothstep(years, 2020 - BLEND_HALFWIDTH_AB, 2020 + BLEND_HALFWIDTH_AB)[None, :]
    val_ab = (1 - w_ab) * f_A + w_ab * f_B
    w_bc = smoothstep(years, 2030 - BLEND_HALFWIDTH_BC, 2030 + BLEND_HALFWIDTH_BC)[None, :]
    val = (1 - w_bc) * val_ab + w_bc * f_C
    val = np.clip(val, 0.0, None)   # negative wire length/copper is unphysical

    return CategoryDraw(val, u, start_fraction, target_fraction, tau)


# --------------------------------------------------------------------------
# 4. MONTE CARLO ENGINE
# --------------------------------------------------------------------------

@dataclass
class MCResult:
    years: np.ndarray
    segments: list
    metrics: list
    categories: list
    raw_total: dict            # raw_total[metric][segment] -> (N_ITER, n_years)
    raw_cat: dict = field(default_factory=dict)       # (wt,seg,metric) -> (N_ITER, n_years)
    draws: dict = field(default_factory=dict)         # (wt,seg,metric) -> CategoryDraw
    raw_catgroup: dict = field(default_factory=dict)  # (wire_category,seg,metric) -> (N_ITER, n_years)
    wire_category_names: list = field(default_factory=list)  # e.g. HV Power Cable, Coaxial Cable, ...


def run_monte_carlo(n_iter=N_ITER, start_year=START_YEAR, end_year=END_YEAR,
                     seed=RANDOM_SEED) -> MCResult:
    rng = np.random.default_rng(seed)
    wire_df = load_wire_level_data()
    categories = build_category_anchor_table(wire_df)
    years = np.arange(start_year, end_year + 1)
    n_years = len(years)

    all_segments = BASE_SEGMENTS + list(J_SUFFIX.values())
    raw_total = {m: {s: np.zeros((n_iter, n_years)) for s in all_segments} for m in METRICS}
    raw_cat, draws = {}, {}

    # ONE trim percentile per segment per iteration, shared by all 28 wire
    # types in that segment. This is the correlation assumption documented at
    # TRIM_JITTER_STD: iteration i represents ONE simulated vehicle, wired
    # consistently sparse or consistently rich across its whole harness.
    # UNIT: dimensionless percentile in [0,1].
    u_trim = {seg: rng.uniform(0, 1, size=n_iter) for seg in BASE_SEGMENTS}

    for cat in categories:
        d = build_category_trajectories(cat, years, rng, u_trim[cat.segment], n_iter)
        raw_total[cat.metric][cat.segment] += d.traj
        key = (cat.wire_type, cat.segment, cat.metric)
        raw_cat[key] = d.traj
        draws[key] = d

    for seg in BASE_SEGMENTS:
        height_adder = sample_triangular_tri(rng, HEIGHT_ADDER_TRI, n_iter)[:, None]
        for metric in METRICS:
            raw_total[metric][J_SUFFIX[seg]] = raw_total[metric][seg] * (1 + height_adder)

    # --- wire-category-group aggregation: sum all wire types sharing a
    #     Wire_Category (e.g. "HV Power Cable", "Coaxial Cable", ...), per
    #     segment and metric. Base segments: straight sum of raw_cat.
    #     J* segments: same height-adder scaling applied to the base
    #     segment's category sum (consistent with how raw_total handles J*).
    wire_category_names = sorted({c.wire_category for c in categories})
    raw_catgroup = {}
    for wcat in wire_category_names:
        for metric in METRICS:
            for seg in BASE_SEGMENTS:
                members = [c for c in categories
                           if c.wire_category == wcat and c.segment == seg and c.metric == metric]
                total = np.zeros((n_iter, n_years))
                for c in members:
                    total += raw_cat[(c.wire_type, c.segment, c.metric)]
                raw_catgroup[(wcat, seg, metric)] = total
            # J* segment: scale the base segment's category sum by that
            # segment's own height-adder draw (re-derive deterministically
            # is not possible since height_adder wasn't stored -- instead
            # scale by the ratio actually realized in raw_total, which is
            # exact and avoids re-sampling).
    for wcat in wire_category_names:
        for metric in METRICS:
            for base_seg, j_seg in J_SUFFIX.items():
                base_total = raw_total[metric][base_seg]
                j_total = raw_total[metric][j_seg]
                with np.errstate(divide="ignore", invalid="ignore"):
                    ratio = np.where(base_total > 1e-12, j_total / base_total, 1.0)
                raw_catgroup[(wcat, j_seg, metric)] = raw_catgroup[(wcat, base_seg, metric)] * ratio

    return MCResult(years, all_segments, METRICS, categories, raw_total, raw_cat, draws,
                     raw_catgroup, wire_category_names)


# --------------------------------------------------------------------------
# 5. STATS + HISTOGRAMS (stock-and-flow-compatible schema)
# --------------------------------------------------------------------------

def hist_mode(arr: np.ndarray, n_bins: int = N_HIST_BINS):
    """Mode estimated as the CENTRE OF THE FULLEST HISTOGRAM BIN.

    Deliberately not a KDE: it uses the same n_bins as the exported histogram
    CSV, so the reported Mode can be reproduced exactly from that file. The
    price is that Mode is quantised to the bin width and will shift slightly
    if N_HIST_BINS changes.

    Returns:
        (mode value, counts, bin edges) -- the latter two so callers that also
        want the histogram do not have to recompute it.
    """
    counts, edges = np.histogram(arr, bins=n_bins)
    i = np.argmax(counts)
    return 0.5 * (edges[i] + edges[i + 1]), counts, edges


def build_stats_table(result: MCResult) -> pd.DataFrame:
    """Full time series (every year), three levels stacked in one tidy table
    (filter on `Level`): 'WireType' (28 types), 'Category' (7 groups: HV
    Power Cable, HV Safety Wire, LV Power Cable, Signal Wire, Sensor Wire,
    Coaxial Cable, Safety Wire), 'Total' (whole segment)."""
    rows = []
    for cat in result.categories:
        if cat.cls == "absent":
            continue   # excluded from output -- reported once in the run log instead
        arr = result.raw_cat[(cat.wire_type, cat.segment, cat.metric)]  # (N_ITER, n_years)
        p2_5, median, p97_5 = np.percentile(arr, [2.5, 50, 97.5], axis=0)
        mean = arr.mean(axis=0)
        for i, yr in enumerate(result.years):
            mode_val, _, _ = hist_mode(arr[:, i])
            rows.append(dict(
                Level="WireType", Segment=cat.segment, Drivetrain=DRIVETRAIN, Wire_Type=cat.wire_type,
                Wire_Category=cat.wire_category, Class=cat.cls, Metric=cat.metric,
                Year=int(yr), N_Iter=arr.shape[0],
                Mean=mean[i], P2_5=p2_5[i], Median=median[i], Mode=mode_val, P97_5=p97_5[i],
            ))
    # category-group summaries (e.g. all HV Power Cable types summed, all
    # Coaxial Cable types summed, ...) -- base AND J* segments
    for metric in result.metrics:
        for seg in result.segments:
            for wcat in result.wire_category_names:
                arr = result.raw_catgroup[(wcat, seg, metric)]
                p2_5, median, p97_5 = np.percentile(arr, [2.5, 50, 97.5], axis=0)
                mean = arr.mean(axis=0)
                for i, yr in enumerate(result.years):
                    mode_val, _, _ = hist_mode(arr[:, i])
                    rows.append(dict(
                        Level="Category", Segment=seg, Drivetrain=DRIVETRAIN, Wire_Type=wcat,
                        Wire_Category=wcat, Class="", Metric=metric,
                        Year=int(yr), N_Iter=arr.shape[0],
                        Mean=mean[i], P2_5=p2_5[i], Median=median[i], Mode=mode_val, P97_5=p97_5[i],
                    ))
    # segment totals (sum of all 28 wire types -- BASE and J* segments)
    for metric in result.metrics:
        for seg in result.segments:
            arr = result.raw_total[metric][seg]
            p2_5, median, p97_5 = np.percentile(arr, [2.5, 50, 97.5], axis=0)
            mean = arr.mean(axis=0)
            for i, yr in enumerate(result.years):
                mode_val, _, _ = hist_mode(arr[:, i])
                rows.append(dict(
                    Level="Total", Segment=seg, Drivetrain=DRIVETRAIN, Wire_Type="TOTAL",
                    Wire_Category="TOTAL", Class="", Metric=metric,
                    Year=int(yr), N_Iter=arr.shape[0],
                    Mean=mean[i], P2_5=p2_5[i], Median=median[i], Mode=mode_val, P97_5=p97_5[i],
                ))
    return pd.DataFrame(rows)


def build_histogram_table(result: MCResult, years=SNAPSHOT_YEARS) -> pd.DataFrame:
    rows = []
    year_idx = {yr: np.searchsorted(result.years, yr) for yr in years}

    def add(level, seg, wt, wcat, metric, arr):
        for yr, i in year_idx.items():
            counts, edges = np.histogram(arr[:, i], bins=N_HIST_BINS)
            for b in range(N_HIST_BINS):
                rows.append(dict(Level=level, Segment=seg, Drivetrain=DRIVETRAIN, Wire_Type=wt,
                                  Wire_Category=wcat, Metric=metric, Year=yr,
                                  Bin_Left=edges[b], Bin_Right=edges[b + 1], Count=int(counts[b])))

    for cat in result.categories:
        if cat.cls == "absent":
            continue   # excluded from output -- reported once in the run log instead
        arr = result.raw_cat[(cat.wire_type, cat.segment, cat.metric)]
        add("WireType", cat.segment, cat.wire_type, cat.wire_category, cat.metric, arr)
    for metric in result.metrics:
        for seg in result.segments:
            for wcat in result.wire_category_names:
                add("Category", seg, wcat, wcat, metric, result.raw_catgroup[(wcat, seg, metric)])
    for metric in result.metrics:
        for seg in result.segments:
            add("Total", seg, "TOTAL", "TOTAL", metric, result.raw_total[metric][seg])

    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 6. SENSITIVITY (Spearman, signed, sorted by |rho|)
# --------------------------------------------------------------------------

def param_tornado_data(result: MCResult, wire_type, segment, metric, year) -> pd.DataFrame:
    d = result.draws[(wire_type, segment, metric)]
    idx = np.searchsorted(result.years, year)
    y = result.raw_cat[(wire_type, segment, metric)][:, idx]
    params = {"trim_percentile": d.u_eff, "start_fraction": d.start_fraction,
              "target_fraction": d.target_fraction, "decay_tau": d.tau}
    rows = []
    for name, x in params.items():
        if np.std(x) < 1e-12 or np.std(y) < 1e-12:
            rho = 0.0
        else:
            rho, _ = spearmanr(x, y)
        rows.append(dict(Wire_Type=wire_type, Segment=segment, Metric=metric, Year=year,
                          Parameter=name, Spearman_rho=rho))
    df = pd.DataFrame(rows)
    return df.reindex(df.Spearman_rho.abs().sort_values(ascending=True).index)


def contribution_tornado_data(result: MCResult, segment, metric, year) -> pd.DataFrame:
    idx = np.searchsorted(result.years, year)
    y = result.raw_total[metric][segment][:, idx]
    rows = []
    for cat in result.categories:
        if cat.segment != segment or cat.metric != metric or cat.cls == "absent":
            continue
        x = result.raw_cat[(cat.wire_type, segment, metric)][:, idx]
        if np.std(x) < 1e-12 or np.std(y) < 1e-12:
            rho = 0.0
        else:
            rho, _ = spearmanr(x, y)
        rows.append(dict(Wire_Type=cat.wire_type, Segment=segment, Metric=metric, Year=year,
                          Spearman_rho=rho))
    df = pd.DataFrame(rows)
    return df.reindex(df.Spearman_rho.abs().sort_values(ascending=True).index)


def _tornado_ax(ax, labels, rhos, title):
    colors = ["#d62728" if r < 0 else "#1f77b4" for r in rhos]
    ax.barh(labels, rhos, color=colors, height=0.7)
    ax.axvline(0, color="k", lw=0.8)
    ax.set_xlim(-1, 1)
    ax.set_title(title, fontsize=8)
    ax.tick_params(labelsize=6)


# --------------------------------------------------------------------------
# 7. PLOTTING
# --------------------------------------------------------------------------

def plot_category_grid(result: MCResult, segment, metric, out_path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    cats = sorted({c.wire_type for c in result.categories
                    if c.segment == segment and c.metric == metric and c.cls != "absent"})
    ncols, nrows = 4, int(np.ceil(len(cats) / 4))
    fig, axes = plt.subplots(nrows, ncols, figsize=(16, 3.0 * nrows), sharex=True)
    axes = axes.flatten()
    for ax, wt in zip(axes, cats):
        arr = result.raw_cat[(wt, segment, metric)]
        p2_5, median, p97_5 = np.percentile(arr, [2.5, 50, 97.5], axis=0)
        ax.plot(result.years, median, color="#1f77b4", lw=1.3)
        ax.fill_between(result.years, p2_5, p97_5, color="#1f77b4", alpha=0.2)
        ax.axvline(2020, color="gray", lw=0.5, ls=":")
        ax.axvline(2030, color="gray", lw=0.5, ls=":")
        cls = next(c.cls for c in result.categories if c.wire_type == wt and c.segment == segment)
        ax.set_title(f"{wt} [{cls}]", fontsize=8)
        ax.tick_params(labelsize=6)
    for ax in axes[len(cats):]:
        ax.axis("off")
    fig.suptitle(f"{segment} ({DRIVETRAIN}) -- {metric}, per wire category (median, P2.5-P97.5)", fontsize=11)
    plt.tight_layout(rect=[0, 0, 1, 0.97])
    plt.savefig(out_path, dpi=130)
    plt.close(fig)


def plot_histogram_grid(result: MCResult, segment, metric, year, out_path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    idx = np.searchsorted(result.years, year)
    cats = sorted({c.wire_type for c in result.categories
                    if c.segment == segment and c.metric == metric and c.cls != "absent"})
    ncols, nrows = 4, int(np.ceil(len(cats) / 4))
    fig, axes = plt.subplots(nrows, ncols, figsize=(16, 3.0 * nrows))
    axes = axes.flatten()
    for ax, wt in zip(axes, cats):
        arr = result.raw_cat[(wt, segment, metric)][:, idx]
        ax.hist(arr, bins=N_HIST_BINS, color="#1f77b4", alpha=0.8)
        ax.set_title(wt, fontsize=8)
        ax.tick_params(labelsize=6)
    for ax in axes[len(cats):]:
        ax.axis("off")
    fig.suptitle(f"{segment} ({DRIVETRAIN}) -- {metric} distribution in {year} (50-bin histograms)", fontsize=11)
    plt.tight_layout(rect=[0, 0, 1, 0.97])
    plt.savefig(out_path, dpi=130)
    plt.close(fig)


# --- wire-CATEGORY-GROUP level (7 groups: HV Power Cable, HV Safety Wire,
#     LV Power Cable, Signal Wire, Sensor Wire, Coaxial Cable, Safety Wire)
#     -- sits between the 28 individual wire types and the single segment
#     total. ---

def plot_wirecategory_grid(result: MCResult, segment, metric, out_path):
    """Separated (not combined) small-multiples grid, one axes per
    wire-category group, analogous to plot_category_grid but one level up."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    wcats = result.wire_category_names
    ncols, nrows = 4, int(np.ceil(len(wcats) / 4))
    fig, axes = plt.subplots(nrows, ncols, figsize=(16, 3.2 * nrows), sharex=True)
    axes = np.atleast_1d(axes).flatten()
    for ax, wcat in zip(axes, wcats):
        arr = result.raw_catgroup[(wcat, segment, metric)]
        p2_5, median, p97_5 = np.percentile(arr, [2.5, 50, 97.5], axis=0)
        ax.plot(result.years, median, color="#d62728", lw=1.6)
        ax.fill_between(result.years, p2_5, p97_5, color="#d62728", alpha=0.18)
        ax.axvline(2020, color="gray", lw=0.5, ls=":")
        ax.axvline(2030, color="gray", lw=0.5, ls=":")
        ax.set_title(wcat, fontsize=9)
        ax.tick_params(labelsize=7)
    for ax in axes[len(wcats):]:
        ax.axis("off")
    fig.suptitle(f"{segment} ({DRIVETRAIN}) -- {metric}, summed by wire-category group "
                 f"(median, P2.5-P97.5)", fontsize=11)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.savefig(out_path, dpi=130)
    plt.close(fig)


def plot_wirecategory_histogram_grid(result: MCResult, segment, metric, year, out_path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    idx = np.searchsorted(result.years, year)
    wcats = result.wire_category_names
    ncols, nrows = 4, int(np.ceil(len(wcats) / 4))
    fig, axes = plt.subplots(nrows, ncols, figsize=(16, 3.2 * nrows))
    axes = np.atleast_1d(axes).flatten()
    for ax, wcat in zip(axes, wcats):
        arr = result.raw_catgroup[(wcat, segment, metric)][:, idx]
        ax.hist(arr, bins=N_HIST_BINS, color="#d62728", alpha=0.8)
        ax.set_title(wcat, fontsize=9)
        ax.tick_params(labelsize=7)
    for ax in axes[len(wcats):]:
        ax.axis("off")
    fig.suptitle(f"{segment} ({DRIVETRAIN}) -- {metric} distribution in {year}, "
                 f"by wire-category group (50-bin histograms)", fontsize=11)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.savefig(out_path, dpi=130)
    plt.close(fig)


def plot_category_overview(result: MCResult, segment, metric, out_path):
    """The 7 wire-category groups overlaid as lines (not stacked) for one
    segment, with the segment TOTAL drawn on top for reference -- bridges
    the per-wire-type detail and the single overall-summary total."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(10, 6.5))
    cmap = plt.get_cmap("tab10")
    for i, wcat in enumerate(result.wire_category_names):
        arr = result.raw_catgroup[(wcat, segment, metric)]
        median = np.percentile(arr, 50, axis=0)
        ax.plot(result.years, median, color=cmap(i), lw=1.6, label=wcat)

    total_arr = result.raw_total[metric][segment]
    t_p2_5, t_median, t_p97_5 = np.percentile(total_arr, [2.5, 50, 97.5], axis=0)
    ax.plot(result.years, t_median, color="black", lw=2.4, ls="--", label="TOTAL (segment)")
    ax.fill_between(result.years, t_p2_5, t_p97_5, color="black", alpha=0.08)

    ax.axvline(2020, color="gray", lw=0.7, ls=":")
    ax.axvline(2030, color="gray", lw=0.7, ls=":")
    ax.set_title(f"{segment} ({DRIVETRAIN}) -- {metric}: category-group summary vs. total")
    ax.set_xlabel("Year")
    ax.set_ylabel(metric)
    ax.legend(fontsize=8, ncol=2)
    ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(out_path, dpi=140)
    plt.close(fig)


def plot_param_tornado_grid(result: MCResult, segment, metric, year, out_path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    cats = sorted({c.wire_type for c in result.categories
                    if c.segment == segment and c.cls != "absent"})
    ncols, nrows = 4, int(np.ceil(len(cats) / 4))
    fig, axes = plt.subplots(nrows, ncols, figsize=(16, 3.0 * nrows))
    axes = axes.flatten()
    for ax, wt in zip(axes, cats):
        df = param_tornado_data(result, wt, segment, metric, year)
        _tornado_ax(ax, df.Parameter, df.Spearman_rho, wt)
    for ax in axes[len(cats):]:
        ax.axis("off")
    fig.suptitle(f"{segment} ({DRIVETRAIN}) -- {metric} @ {year}: parameter sensitivity (signed Spearman rho)",
                 fontsize=11)
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    plt.savefig(out_path, dpi=130)
    plt.close(fig)


def plot_contribution_tornado(result: MCResult, segment, metric, out_path, years=TORNADO_YEARS):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, axes = plt.subplots(1, len(years), figsize=(6 * len(years), 8), sharey=False)
    for ax, year in zip(axes, years):
        df = contribution_tornado_data(result, segment, metric, year)
        _tornado_ax(ax, df.Wire_Type, df.Spearman_rho, f"{year}")
        ax.tick_params(labelsize=7)
    fig.suptitle(f"{segment} ({DRIVETRAIN}) -- which wire categories drive TOTAL {metric} uncertainty?",
                 fontsize=11)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.savefig(out_path, dpi=130)
    plt.close(fig)


def plot_totals_overview(result: MCResult, stats: pd.DataFrame, out_path: str):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    colors = {"AB": "#1f77b4", "CD": "#ff7f0e", "EF": "#2ca02c",
              "JA-JB": "#7fb3e0", "JC-JD": "#ffb877", "JE-JF": "#8fd18f"}
    seg_order = ["AB", "JA-JB", "CD", "JC-JD", "EF", "JE-JF"]
    tot = stats[stats.Wire_Type == "TOTAL"]

    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5), sharex=True)
    for ax, metric, title, unit in zip(
        axes, ["Cu (kg)", "Length (m)"],
        ["Copper mass per vehicle (TOTAL, all categories)",
         "Wire length per vehicle (TOTAL, all categories)"], ["kg", "m"],
    ):
        sub = tot[tot.Metric == metric]
        for seg in seg_order:
            s = sub[sub.Segment == seg].sort_values("Year")
            ls = "--" if seg.startswith("J") else "-"
            ax.plot(s.Year, s.Median, ls, color=colors[seg], label=seg, linewidth=1.8)
            if not seg.startswith("J"):
                ax.fill_between(s.Year, s.P2_5, s.P97_5, color=colors[seg], alpha=0.12)
        ax.axvline(2020, color="gray", lw=0.7, ls=":")
        ax.axvline(2030, color="gray", lw=0.7, ls=":")
        ax.set_title(title); ax.set_ylabel(unit); ax.set_xlabel("Year"); ax.grid(alpha=0.3)
    axes[0].legend(fontsize=8, ncol=2)
    plt.tight_layout()
    plt.savefig(out_path, dpi=140)
    plt.close(fig)


# --------------------------------------------------------------------------
# 8. MAIN
# --------------------------------------------------------------------------

if __name__ == "__main__":
    print("Running Monte Carlo...")
    result = run_monte_carlo()

    absent = sorted({(c.wire_type, c.segment) for c in result.categories if c.cls == "absent"})
    print(f"\nExcluded {len(absent)} wire-type/segment combo(s) classified 'absent' "
          f"(zero in both Historical and Future data -- e.g. rear-motor cabling on a "
          f"single-motor segment). They are omitted from all stats, histograms, and "
          f"plots below; their contribution to category/segment totals is 0 by construction.")
    for wt, seg in absent:
        print(f"    absent: {wt} / {seg}")

    print("\nBuilding stats table...")
    stats = build_stats_table(result)
    stats.to_csv(OUT_SUBDIRS["data"] / "bev_wiring_stats.csv", index=False)
    print(f"  {len(stats)} rows -> data/bev_wiring_stats.csv")

    print("Building histogram table...")
    hist = build_histogram_table(result)
    hist.to_csv(OUT_SUBDIRS["data"] / "bev_wiring_histograms.csv", index=False)
    print(f"  {len(hist)} rows -> data/bev_wiring_histograms.csv")

    print("Segment totals overview (all 6 segments, incl. J*)...")
    plot_totals_overview(result, stats, OUT_SUBDIRS["totals"] / "totals_overview.png")

    print("Category trajectory grids (separated, not combined; base segments only --")
    print("  J* segments are a fixed +height% scaling of their base segment's categories)...")
    for seg in BASE_SEGMENTS:
        for metric in result.metrics:
            plot_category_grid(result, seg, metric,
                                OUT_SUBDIRS["trajectories"] / f"grid_{seg}_{METRIC_SLUG[metric]}.png")

    print("Histogram grids (2070 snapshot, per segment/metric)...")
    for seg in BASE_SEGMENTS:
        for metric in result.metrics:
            plot_histogram_grid(result, seg, metric, 2070,
                                 OUT_SUBDIRS["histograms"] / f"hist_{seg}_{METRIC_SLUG[metric]}_2070.png")

    print("Wire-category-group trajectory grids (HV / Coax / LV / Signal / Sensor / Safety...)...")
    for seg in BASE_SEGMENTS:
        for metric in result.metrics:
            plot_wirecategory_grid(result, seg, metric,
                                    OUT_SUBDIRS["category_trajectories"] / f"catgrid_{seg}_{METRIC_SLUG[metric]}.png")

    print("Wire-category-group histogram grids (2070 snapshot)...")
    for seg in BASE_SEGMENTS:
        for metric in result.metrics:
            plot_wirecategory_histogram_grid(result, seg, metric, 2070,
                                              OUT_SUBDIRS["category_histograms"] / f"cathist_{seg}_{METRIC_SLUG[metric]}_2070.png")

    print("Category-group vs. total overview (all 6 segments, incl. J*)...")
    for seg in result.segments:
        for metric in result.metrics:
            plot_category_overview(result, seg, metric,
                                    OUT_SUBDIRS["category_overview"] / f"catoverview_{seg}_{METRIC_SLUG[metric]}.png")

    print("Parameter-sensitivity tornado grids (2070, base segments)...")
    for seg in BASE_SEGMENTS:
        for metric in result.metrics:
            plot_param_tornado_grid(result, seg, metric, 2070,
                                     OUT_SUBDIRS["sensitivity"] / f"tornado_params_{seg}_{METRIC_SLUG[metric]}_2070.png")

    print("Contribution-sensitivity tornado plots (2030/2050/2070, base segments)...")
    for seg in BASE_SEGMENTS:
        for metric in result.metrics:
            plot_contribution_tornado(result, seg, metric,
                                       OUT_SUBDIRS["sensitivity"] / f"tornado_contrib_{seg}_{METRIC_SLUG[metric]}.png")

    print("\nDone. Output layout:")
    print(f"  {OUT_SUBDIRS['data']}")
    print(f"  {OUT_SUBDIRS['totals']}")
    print(f"  {OUT_SUBDIRS['trajectories']}")
    print(f"  {OUT_SUBDIRS['histograms']}")
    print(f"  {OUT_SUBDIRS['category_trajectories']}")
    print(f"  {OUT_SUBDIRS['category_histograms']}")
    print(f"  {OUT_SUBDIRS['category_overview']}")
    print(f"  {OUT_SUBDIRS['sensitivity']}")

    print("\nSegment TOTAL median Cu (kg) sanity check:")
    tot = stats[(stats.Wire_Type == "TOTAL") & (stats.Metric == "Cu (kg)")]
    for seg in result.segments:
        s = tot[tot.Segment == seg].set_index("Year")
        vals = [f"{s.loc[y, 'Median']:5.1f}" for y in [2010, 2020, 2025, 2030, 2040, 2050, 2070]]
        print(f"  {seg:8s}: " + "  ".join(vals))